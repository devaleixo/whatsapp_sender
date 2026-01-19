#!/usr/bin/env python3
"""
WhatsApp Sender
Envia mensagens WhatsApp para contatos de um arquivo XLSX (gerado pelo Google Scraper)
"""

import os
import sys
import time
import re
from typing import Optional
from dataclasses import dataclass

from openpyxl import load_workbook
from evolution_client import EvolutionAPI


@dataclass
class Contact:
    """Representa um contato para envio"""
    nome: str
    telefone: str
    endereco: Optional[str] = None
    avaliacao: Optional[str] = None
    website: Optional[str] = None


class WhatsAppSender:
    """Envia mensagens WhatsApp em massa"""
    
    def __init__(self, instance_name: str = "business_sender",
                 api_url: str = "http://localhost:8080",
                 api_key: str = "whatsapp_sender_secret_key_2024"):
        self.instance_name = instance_name
        self.api = EvolutionAPI(api_url, api_key)
        self.sent_count = 0
        self.failed_count = 0
        self.skipped_count = 0
    
    def setup(self) -> bool:
        """Configura a instância e verifica conexão"""
        print("🔧 Configurando WhatsApp Sender...")
        
        # Verifica se a instância existe
        instances = self.api.list_instances()
        instance_exists = False
        
        if isinstance(instances, list):
            for inst in instances:
                if inst.get("name") == self.instance_name:
                    instance_exists = True
                    break
        
        if not instance_exists:
            print(f"   Criando instância '{self.instance_name}'...")
            result = self.api.create_instance(self.instance_name)
            # Ignora erro se a instância já existe
            if result.get("error"):
                error_msg = str(result.get("message", ""))
                if "already in use" in error_msg or "already exists" in error_msg:
                    print(f"   ✓ Instância '{self.instance_name}' já existe!")
                    instance_exists = True
                else:
                    print(f"   ❌ Erro ao criar instância: {error_msg}")
                    return False
            else:
                print("   ✓ Instância criada!")
        
        # Verifica se está conectado
        if self.api.is_connected(self.instance_name):
            print("   ✓ WhatsApp já está conectado!")
            return True
        
        # Obtém QR Code (com retry, pois pode demorar para gerar)
        print("\n📱 Obtendo QR Code...")
        qr = None
        qr_displayed = False
        
        for attempt in range(5):
            qr = self.api.get_qrcode(self.instance_name)
            if qr.get("code") or qr.get("base64"):
                break
            print(f"   Aguardando geração do QR... ({attempt + 1}/5)")
            time.sleep(2)
        
        # Tenta exibir QR Code no terminal
        if qr and qr.get("code"):
            try:
                import qrcode
                qr_obj = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=1,
                    border=1,
                )
                qr_obj.add_data(qr.get("code"))
                qr_obj.make(fit=True)
                print("\n" + "=" * 50)
                qr_obj.print_ascii(invert=True)
                print("=" * 50)
                qr_displayed = True
            except ImportError:
                print("   ⚠️  Instale 'qrcode' para ver no terminal: pip install qrcode")
        
        # Salva QR Code como imagem
        if qr and qr.get("base64"):
            import base64
            import subprocess
            qr_data = qr.get("base64").split(",")[-1]
            qr_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qrcode.png")
            with open(qr_path, "wb") as f:
                f.write(base64.b64decode(qr_data))
            print(f"\n   💾 QR Code salvo em: {qr_path}")
            
            # Tenta abrir a imagem automaticamente
            try:
                if sys.platform == "linux":
                    subprocess.Popen(["xdg-open", qr_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", qr_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                elif sys.platform == "win32":
                    os.startfile(qr_path)
                print("   📲 Imagem aberta - escaneie com WhatsApp > Aparelhos Conectados")
            except Exception:
                print("   📲 Abra a imagem manualmente e escaneie com WhatsApp")
        
        # Se não conseguiu QR Code, abre o Manager Web
        if not qr_displayed and not (qr and qr.get("base64")):
            import subprocess
            import webbrowser
            manager_url = "http://localhost:8080/manager"
            print(f"\n   ⚠️  QR Code não disponível via API")
            print(f"\n   🌐 Abrindo Manager Web: {manager_url}")
            print(f"   🔑 API Key: whatsapp_sender_secret_key_2024")
            print(f"   📱 Clique na instância '{self.instance_name}' e escaneie o QR")
            try:
                webbrowser.open(manager_url)
            except Exception:
                pass
        
        # Aguarda conexão
        print("\n⏳ Aguardando conexão (3 minutos)...")
        print("   Escaneie o QR Code com WhatsApp > Aparelhos Conectados")
        if self.api.wait_for_connection(self.instance_name, timeout=180):
            print("   ✓ WhatsApp conectado com sucesso!")
            return True
        else:
            print("   ❌ Timeout - QR Code não foi escaneado")
            return False
    
    def load_contacts_from_xlsx(self, filepath: str) -> list[Contact]:
        """Carrega contatos de um arquivo XLSX"""
        if not os.path.exists(filepath):
            print(f"❌ Arquivo não encontrado: {filepath}")
            return []
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        contacts = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Pula cabeçalho
            nome = row[0] if len(row) > 0 else None
            telefone = row[1] if len(row) > 1 else None
            endereco = row[2] if len(row) > 2 else None
            avaliacao = row[3] if len(row) > 3 else None
            website = row[4] if len(row) > 4 else None
            
            # Só adiciona se tiver nome e telefone válido
            if nome and telefone and telefone != "N/A":
                contacts.append(Contact(
                    nome=nome,
                    telefone=str(telefone),
                    endereco=endereco,
                    avaliacao=str(avaliacao) if avaliacao else None,
                    website=website
                ))
        
        return contacts
    
    def format_message(self, template: str, contact: Contact) -> str:
        """Formata a mensagem com os dados do contato"""
        message = template
        message = message.replace("{nome}", contact.nome)
        message = message.replace("{telefone}", contact.telefone)
        message = message.replace("{endereco}", contact.endereco or "")
        message = message.replace("{avaliacao}", contact.avaliacao or "")
        message = message.replace("{website}", contact.website or "")
        return message
    
    def send_messages(self, contacts: list[Contact], message_template: str,
                      delay_seconds: float = 5.0, verify_whatsapp: bool = True) -> dict:
        """
        Envia mensagens para todos os contatos
        
        Args:
            contacts: Lista de contatos
            message_template: Template da mensagem (use {nome}, {endereco}, etc)
            delay_seconds: Atraso entre mensagens (evita bloqueio)
            verify_whatsapp: Se True, verifica se o número tem WhatsApp antes
        
        Returns:
            Resumo do envio
        """
        total = len(contacts)
        print(f"\n📤 Iniciando envio para {total} contatos...")
        print(f"   Delay entre mensagens: {delay_seconds}s")
        print(f"   Verificar WhatsApp: {'Sim' if verify_whatsapp else 'Não'}")
        print("-" * 50)
        
        for i, contact in enumerate(contacts, 1):
            print(f"\n[{i}/{total}] {contact.nome[:40]}...")
            
            # Verifica se tem WhatsApp
            if verify_whatsapp:
                if not self.api.has_whatsapp(self.instance_name, contact.telefone):
                    print(f"   ⚠️  Sem WhatsApp: {contact.telefone}")
                    self.skipped_count += 1
                    continue
            
            # Formata e envia mensagem
            message = self.format_message(message_template, contact)
            result = self.api.send_text(self.instance_name, contact.telefone, message)
            
            if result.get("error"):
                print(f"   ❌ Erro: {result.get('message', 'Desconhecido')[:50]}")
                self.failed_count += 1
            else:
                print(f"   ✓ Enviado para {contact.telefone}")
                self.sent_count += 1
            
            # Delay para evitar bloqueio
            if i < total:
                time.sleep(delay_seconds)
        
        # Resumo
        print("\n" + "=" * 50)
        print("📊 RESUMO DO ENVIO")
        print("=" * 50)
        print(f"   ✓ Enviados:  {self.sent_count}")
        print(f"   ❌ Erros:     {self.failed_count}")
        print(f"   ⚠️  Pulados:   {self.skipped_count}")
        print(f"   📋 Total:     {total}")
        
        return {
            "sent": self.sent_count,
            "failed": self.failed_count,
            "skipped": self.skipped_count,
            "total": total
        }


def main():
    """Função principal"""
    if len(sys.argv) < 2:
        print("""
╔═══════════════════════════════════════════════════════════════╗
║              WhatsApp Sender - Evolution API                  ║
╠═══════════════════════════════════════════════════════════════╣
║  Uso:                                                         ║
║    python3 whatsapp_sender.py <arquivo.xlsx> [mensagem]       ║
║                                                               ║
║  Exemplos:                                                    ║
║    python3 whatsapp_sender.py contatos.xlsx                   ║
║    python3 whatsapp_sender.py contatos.xlsx "Olá {nome}!"     ║
║                                                               ║
║  Variáveis na mensagem:                                       ║
║    {nome}      - Nome do negócio                              ║
║    {telefone}  - Telefone                                     ║
║    {endereco}  - Endereço                                     ║
║    {avaliacao} - Nota no Google                               ║
║    {website}   - Site                                         ║
║                                                               ║
║  Configuração:                                                ║
║    1. Inicie o Docker: docker-compose up -d                   ║
║    2. Execute este script                                     ║
║    3. Escaneie o QR Code com seu WhatsApp                     ║
╚═══════════════════════════════════════════════════════════════╝
        """)
        sys.exit(1)
    
    xlsx_file = sys.argv[1]
    
    # Mensagem padrão ou customizada
    if len(sys.argv) > 2:
        message_template = ' '.join(sys.argv[2:])
    else:
        message_template = """Olá! Tudo bem?

Sou desenvolvedor de sites e sistemas. Vi que a empresa *{nome}* pode se beneficiar de uma presença digital profissional.

Ofereço:
✅ Sites modernos e responsivos
✅ Sistemas de agendamento
✅ Integração com WhatsApp
✅ SEO para aparecer no Google

Gostaria de saber mais? Responda esta mensagem!

_Mensagem enviada via sistema automatizado_"""
    
    sender = WhatsAppSender()
    
    # Configura e conecta
    if not sender.setup():
        print("\n❌ Falha ao configurar. Verifique se o Docker está rodando.")
        print("   Execute: docker-compose up -d")
        sys.exit(1)
    
    # Carrega contatos
    print(f"\n📂 Carregando contatos de: {xlsx_file}")
    contacts = sender.load_contacts_from_xlsx(xlsx_file)
    
    if not contacts:
        print("❌ Nenhum contato válido encontrado no arquivo.")
        sys.exit(1)
    
    print(f"   ✓ {len(contacts)} contatos com telefone válido")
    
    # Confirmação
    print(f"\n📝 Mensagem que será enviada:")
    print("-" * 40)
    sample_contact = contacts[0]
    print(sender.format_message(message_template, sample_contact))
    print("-" * 40)
    
    # Verifica flag -y para pular confirmação
    skip_confirm = "-y" in sys.argv or "--yes" in sys.argv
    
    if not skip_confirm:
        confirm = input("\n⚠️  Deseja continuar? (s/N): ").strip().lower()
        if confirm != 's':
            print("Operação cancelada.")
            sys.exit(0)
    else:
        print("\n⚡ Confirmação pulada (-y). Iniciando envio...")
    
    # Envia mensagens
    sender.send_messages(contacts, message_template, delay_seconds=5.0, verify_whatsapp=True)


if __name__ == "__main__":
    main()
